import os
from fpdf import FPDF
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DOWNLOADS_DIR = "/Users/petercurrey/Desktop/Drawdown Trading/public/downloads"
os.makedirs(DOWNLOADS_DIR, exist_ok=True)

class StyledPDF(FPDF):
    def header(self):
        self.set_fill_color(8, 10, 15) # Dark header
        self.rect(0, 0, 210, 15, 'F')
        self.set_font("Helvetica", "B", 9)
        self.set_text_color(200, 241, 53) # Neon #C8F135
        self.set_xy(10, 3)
        self.cell(0, 10, "DRAWDOWN TRADING // INSTITUTIONAL PLAYBOOK SERIES")
        self.set_text_color(180, 180, 180)
        self.cell(0, 10, "DRAWDOWN.TRADING", align="R")
        self.ln(12)

    def footer(self):
        self.set_y(-15)
        self.set_font("Helvetica", "I", 8)
        self.set_text_color(120, 120, 120)
        self.cell(0, 10, f"Page {self.page_no()} | Confidential & Proprietary - Drawdown Trading", align="C")

# 1. Generate risk-management-guide.pdf
pdf1 = StyledPDF()
pdf1.add_page()
pdf1.set_auto_page_break(auto=True, margin=15)

pdf1.set_font("Helvetica", "B", 20)
pdf1.set_text_color(20, 20, 20)
pdf1.cell(0, 12, "Drawdown Risk Management Guide")
pdf1.ln(10)

pdf1.set_font("Helvetica", "B", 11)
pdf1.set_text_color(200, 100, 0)
pdf1.cell(0, 8, "The Mathematical & Execution Blueprint for Capital Preservation")
pdf1.ln(8)

pdf1.set_font("Helvetica", "", 10)
pdf1.set_text_color(40, 40, 40)
pdf1.multi_cell(0, 6, "Risk management is the single determinant of longevity in professional trading. This guide details the exact mathematical boundaries, position sizing formulas, and drawdown mitigation protocols used by institutional desks and top-tier funded traders.")
pdf1.ln(6)

pdf1.set_font("Helvetica", "B", 13)
pdf1.set_text_color(10, 10, 10)
pdf1.cell(0, 8, "1. Core Risk Parameters & Fixed Fractional Sizing")
pdf1.ln(7)

pdf1.set_font("Helvetica", "", 10)
pdf1.multi_cell(0, 6, "- Maximum Risk Per Trade: Never exceed 1.0% of current account equity (0.5% recommended during drawdown phases).\n- Maximum Daily Loss Limit: Hard stop at 2.0% daily drawdown. If triggered, terminal is locked until next session.\n- Maximum Trailing Drawdown Limit: Maintain a minimum 3.0% buffer above mandatory evaluation drawdown thresholds.\n- Kelly Criterion Fractional Sizing: Optimal Risk = (Win Rate * R:R - Loss Rate) / R:R * 0.25 (Quarter-Kelly).")
pdf1.ln(6)

pdf1.set_font("Helvetica", "B", 13)
pdf1.cell(0, 8, "2. Position Sizing Formula Table")
pdf1.ln(7)

# Table Header
pdf1.set_font("Helvetica", "B", 9)
pdf1.set_fill_color(30, 40, 55)
pdf1.set_text_color(255, 255, 255)
pdf1.cell(45, 7, "Account Size", border=1, fill=True)
pdf1.cell(45, 7, "Risk % (0.5%)", border=1, fill=True)
pdf1.cell(45, 7, "Risk % (1.0%)", border=1, fill=True)
pdf1.cell(45, 7, "Max Daily Loss (2.0%)", border=1, fill=True)
pdf1.ln(7)

# Table Data
pdf1.set_font("Helvetica", "", 9)
pdf1.set_text_color(30, 30, 30)
data = [
    ("$10,000", "$50", "$100", "$200"),
    ("$25,000", "$125", "$250", "$500"),
    ("$50,000", "$250", "$500", "$1,000"),
    ("$100,000", "$500", "$1,000", "$2,000"),
    ("$200,000", "$1,000", "$2,000", "$4,000"),
]
for row in data:
    pdf1.cell(45, 7, row[0], border=1)
    pdf1.cell(45, 7, row[1], border=1)
    pdf1.cell(45, 7, row[2], border=1)
    pdf1.cell(45, 7, row[3], border=1)
    pdf1.ln(7)

pdf1.ln(6)
pdf1.set_font("Helvetica", "B", 13)
pdf1.cell(0, 8, "3. Falsification Protocol & Execution Rules")
pdf1.ln(7)
pdf1.set_font("Helvetica", "", 10)
pdf1.multi_cell(0, 6, "Every order intent must possess an explicit invalidation thesis prior to entry:\n1. Technical Invalidation: Price closes past key market structure level on anchor timeframe.\n2. Fundamental Invalidation: Macro news release conflicts with fundamental thesis direction.\n3. Time-Based Invalidation: Trade fails to resolve within designated session window (London/NY close).")

pdf1.output(os.path.join(DOWNLOADS_DIR, "risk-management-guide.pdf"))


# 2. Generate challenge-checklist.pdf
pdf2 = StyledPDF()
pdf2.add_page()

pdf2.set_font("Helvetica", "B", 20)
pdf2.set_text_color(20, 20, 20)
pdf2.cell(0, 12, "30-Day Evaluation Challenge Checklist")
pdf2.ln(10)

pdf2.set_font("Helvetica", "B", 11)
pdf2.set_text_color(0, 120, 200)
pdf2.cell(0, 8, "Step-by-Step Prop Firm Pass & Scaling Roadmap")
pdf2.ln(8)

pdf2.set_font("Helvetica", "", 10)
pdf2.set_text_color(40, 40, 40)
pdf2.multi_cell(0, 6, "Follow this mandatory 30-day checklist during your prop firm evaluation to eliminate emotional impulse trades, enforce risk rules, and guarantee disciplined execution.")
pdf2.ln(6)

checklist_items = [
    ("Phase 1: Pre-Challenge Preparation", [
        "[ ] Read and record exact prop firm rules (Static vs Trailing Drawdown, Weekend holding)",
        "[ ] Set up broker terminal with exact risk calculator lot sizing limits",
        "[ ] Clear daily loss limit buffer in AI Trade Journal (max 2% daily loss)",
        "[ ] Define 2 primary trading setups (e.g. Asia Sweep + NY Expansion)"
    ]),
    ("Phase 2: Week 1 & 2 Execution (Capital Preservation)", [
        "[ ] Risk maximum 0.5% per trade for the first 5 trades until in profit buffer",
        "[ ] Log every trade in AI Trade Journal immediately post-execution",
        "[ ] Do not trade 15 minutes before or after high-impact USD/GBP news releases",
        "[ ] Stop trading for the day if 2 consecutive stop-losses are hit"
    ]),
    ("Phase 3: Week 3 & 4 Target Completion", [
        "[ ] Protect 50% of peak profit buffer - never let a +4% account drawdown back to breakeven",
        "[ ] Maintain minimum 1:2 Risk-to-Reward ratio across all executed setups",
        "[ ] Review weekly DCS (Drawdown Consistency Score) report before session start",
        "[ ] Submit evaluation review upon reaching profit target (+8% or +10%)"
    ])
]

for section_title, items in checklist_items:
    pdf2.set_font("Helvetica", "B", 12)
    pdf2.set_text_color(10, 10, 10)
    pdf2.cell(0, 7, section_title)
    pdf2.ln(6)
    pdf2.set_font("Helvetica", "", 9.5)
    pdf2.set_text_color(40, 40, 40)
    for item in items:
        pdf2.cell(0, 6, item)
        pdf2.ln(5)
    pdf2.ln(4)

pdf2.output(os.path.join(DOWNLOADS_DIR, "challenge-checklist.pdf"))


# 3. Generate prop-firm-comparison-sheet.xlsx
wb1 = openpyxl.Workbook()
ws1 = wb1.active
ws1.title = "Prop Firm Comparison"

ws1.append(["Prop Firm", "Evaluation Type", "Profit Target (Phase 1)", "Max Daily Loss", "Max Overall Drawdown", "Drawdown Type", "Profit Split", "News Trading"])

firm_data = [
    ["FTMO", "2-Step Challenge", "10%", "5%", "10%", "Static Balance", "80% - 90%", "Allowed (Holding Rules)"],
    ["FundedNext", "1-Step / 2-Step", "8% / 10%", "5%", "10%", "Balance-Based", "80% - 95%", "Allowed"],
    ["Funding Pips", "2-Step Evaluation", "8%", "5%", "10%", "Static Equity", "80% - 90%", "Allowed"],
    ["Breakout", "1-Step / 2-Step", "10%", "4%", "8%", "Static Equity", "80% - 90%", "Allowed"],
    ["The 5%ers", "Hyper Growth / High Stakes", "8%", "5%", "10%", "Static Balance", "80% - 100%", "Allowed"],
]

for row in firm_data:
    ws1.append(row)

header_fill = PatternFill(start_color="1A202C", end_color="1A202C", fill_type="solid")
header_font = Font(name="Calibri", size=11, bold=True, color="C8F135")
for col_num in range(1, 9):
    cell = ws1.cell(row=1, column=col_num)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")

for col in ws1.columns:
    max_len = max(len(str(cell.value or '')) for cell in col)
    col_letter = get_column_letter(col[0].column)
    ws1.column_dimensions[col_letter].width = max(max_len + 4, 15)

wb1.save(os.path.join(DOWNLOADS_DIR, "prop-firm-comparison-sheet.xlsx"))


# 4. Generate trading-journal-template.xlsx
wb2 = openpyxl.Workbook()
ws2 = wb2.active
ws2.title = "Trade Journal Template"

ws2.append(["Trade #", "Date", "Instrument", "Direction (Long/Short)", "Entry Price", "Stop Loss", "Take Profit", "Lot Size", "Risk ($)", "P&L ($)", "R:R Multiple", "Setup Notes / Emotion"])

sample_trades = [
    [1, "2026-08-01", "EUR/USD", "Long", 1.08500, 1.08300, 1.08900, 1.50, "$300", "+$600", "2.0R", "Asia liquidity sweep + NY session expansion"],
    [2, "2026-08-01", "GBP/USD", "Short", 1.27800, 1.28100, 1.27200, 1.00, "$300", "+$600", "2.0R", "FCA net short confluence + bearish MSS"],
    [3, "2026-08-02", "XAU/USD", "Long", 2420.00, 2410.00, 2440.00, 0.50, "$500", "-$500", "-1.0R", "Stopped out before FOMC rate decision"],
    [4, "2026-08-02", "USD/JPY", "Short", 154.500, 155.000, 153.500, 1.20, "$600", "+$1,200", "2.0R", "BOJ rate hike rumor + multi-timeframe grid"],
]

for row in sample_trades:
    ws2.append(row)

for col_num in range(1, 13):
    cell = ws2.cell(row=1, column=col_num)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")

for col in ws2.columns:
    max_len = max(len(str(cell.value or '')) for cell in col)
    col_letter = get_column_letter(col[0].column)
    ws2.column_dimensions[col_letter].width = max(max_len + 4, 15)

wb2.save(os.path.join(DOWNLOADS_DIR, "trading-journal-template.xlsx"))

print("ALL 4 DOWNLOAD FILES SUCCESSFULLY GENERATED AS VALID PDF AND EXCEL DOCUMENTS!")
